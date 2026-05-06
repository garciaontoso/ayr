#!/usr/bin/env python3
"""
A&R Monthly Snapshot — genera xlsx mensual con cartera + PnL + dividendos +
earnings updates + thesis scorecards.

Inspirado por skill xlsx-author de Anthropic FSI cookbook.
Coste $0 — corre local con tu propia API key A&R.

Output: ~/Library/Mobile Documents/com~apple~CloudDocs/A&R/snapshots/{YYYY-MM}.xlsx
        (iCloud Drive — accesible desde iPhone, iPad, otros Macs)

Usage:
  python3 scripts/monthly-snapshot.py            # snapshot mes actual
  python3 scripts/monthly-snapshot.py --month 2026-04  # snapshot mes específico

Cron sugerido (1º de cada mes 7am Madrid):
  0 7 1 * * cd /Users/ricardogarciaontoso/IA/AyR && /usr/bin/python3 scripts/monthly-snapshot.py
"""

import os, sys, json, urllib.request, urllib.parse, argparse
from datetime import datetime, date
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip3 install openpyxl")
    sys.exit(1)

API_BASE = "https://api.onto-so.com"
API_TOKEN = os.environ.get("AYR_WORKER_TOKEN") or open(
    Path.home() / ".ayr-env"
).read().split('AYR_WORKER_TOKEN="')[1].split('"')[0]

ICLOUD_DIR = Path.home() / "Library" / "Mobile Documents" / "com~apple~CloudDocs" / "A&R" / "snapshots"


def fetch(path):
    """GET endpoint con auth + Origin headers."""
    url = f"{API_BASE}{path}"
    req = urllib.request.Request(
        url,
        headers={
            "Origin": "https://ayr.onto-so.com",
            "User-Agent": "ayr-monthly-snapshot/1.0",
            "X-AYR-Auth": API_TOKEN,
        },
    )
    with urllib.request.urlopen(req, timeout=30) as r:
        return json.loads(r.read())


def header_style():
    return {
        "font": Font(bold=True, color="FFFFFF", size=10, name="SF Pro Display"),
        "fill": PatternFill("solid", fgColor="C8A44E"),  # A&R gold
        "alignment": Alignment(horizontal="center", vertical="center"),
        "border": Border(bottom=Side(border_style="thin", color="000000")),
    }


def cell_style_money(val):
    if val is None or val == "":
        return {}
    color = "30D158" if val > 0 else "FF453A" if val < 0 else "8E8E93"
    return {"font": Font(color=color, name="SF Mono", size=10, bold=val and abs(val) > 10000)}


def write_header(ws, row, headers, widths=None):
    style = header_style()
    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col_idx, value=h)
        c.font = style["font"]
        c.fill = style["fill"]
        c.alignment = style["alignment"]
        c.border = style["border"]
    if widths:
        for col_idx, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = w


def cover_sheet(wb, snapshot_date):
    ws = wb.create_sheet("Resumen", 0)
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 30

    # Title
    ws.cell(row=1, column=1, value="A&R Monthly Snapshot").font = Font(bold=True, size=22, name="SF Pro Display", color="C8A44E")
    ws.cell(row=2, column=1, value=f"Generado: {snapshot_date.strftime('%d %b %Y')}").font = Font(italic=True, color="8E8E93", size=10)
    ws.cell(row=3, column=1, value="https://ayr.onto-so.com").font = Font(color="0A84FF", size=10)

    # KPIs row
    ws.cell(row=5, column=1, value="MÉTRICAS CLAVE").font = Font(bold=True, size=12)

    return ws


def positions_sheet(wb, positions):
    ws = wb.create_sheet("Cartera")
    headers = ["Ticker", "Empresa", "Sector", "Shares", "Avg Cost", "Last Price", "Market Value USD", "P&L USD", "P&L %", "Yield %", "Weight %"]
    widths = [10, 30, 18, 10, 11, 11, 14, 12, 8, 8, 9]
    write_header(ws, 1, headers, widths)

    total_value = sum(p.get("usd_value", 0) or p.get("market_value", 0) or 0 for p in positions if (p.get("shares", 0) or 0) > 0)

    sorted_pos = sorted([p for p in positions if (p.get("shares", 0) or 0) > 0], key=lambda x: -(x.get("usd_value", 0) or x.get("market_value", 0) or 0))

    row = 2
    for p in sorted_pos:
        shares = p.get("shares", 0) or 0
        if shares <= 0:
            continue
        mv = p.get("usd_value", 0) or p.get("market_value", 0) or 0
        cost = p.get("avg_price", 0) or 0
        cost_total = cost * shares
        pnl_abs = (p.get("pnl_abs") if p.get("pnl_abs") is not None else (mv - cost_total)) or 0
        pnl_pct = (p.get("pnl_pct") if p.get("pnl_pct") is not None else (pnl_abs / cost_total if cost_total > 0 else 0)) or 0
        weight = mv / total_value * 100 if total_value > 0 else 0

        ws.cell(row=row, column=1, value=p.get("ticker", "")).font = Font(bold=True, name="SF Mono", size=10)
        ws.cell(row=row, column=2, value=p.get("name", ""))
        ws.cell(row=row, column=3, value=p.get("sector", ""))
        ws.cell(row=row, column=4, value=shares)
        ws.cell(row=row, column=5, value=cost)
        ws.cell(row=row, column=6, value=p.get("last_price", 0) or 0)
        ws.cell(row=row, column=7, value=mv).number_format = '"$"#,##0'
        cell_pnl = ws.cell(row=row, column=8, value=pnl_abs)
        cell_pnl.number_format = '"$"#,##0;[Red]"-$"#,##0'
        cell_pnl.font = Font(name="SF Mono", size=10, color="30D158" if pnl_abs > 0 else "FF453A" if pnl_abs < 0 else "8E8E93")
        cell_pct = ws.cell(row=row, column=9, value=pnl_pct)
        cell_pct.number_format = "+0.0%;[Red]-0.0%"
        cell_pct.font = Font(name="SF Mono", size=10, color="30D158" if pnl_pct > 0 else "FF453A" if pnl_pct < 0 else "8E8E93")
        ws.cell(row=row, column=10, value=p.get("div_yield", 0) or 0).number_format = "0.0%"
        ws.cell(row=row, column=11, value=weight / 100).number_format = "0.0%"
        row += 1

    # Total row
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=7, value=total_value).number_format = '"$"#,##0'
    ws.cell(row=row, column=7).font = Font(bold=True, name="SF Mono", size=10, color="C8A44E")

    ws.freeze_panes = "A2"
    return total_value, len(sorted_pos)


def pnl_sheet(wb, pnl_data):
    ws = wb.create_sheet("PnL Mensual")
    headers = ["Mes", "Dividendos brutos", "Dividendos netos", "WHT", "Opciones P&L", "Stocks P&L", "TOTAL"]
    widths = [10, 18, 18, 12, 16, 16, 16]
    write_header(ws, 1, headers, widths)

    months_es = ["Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]
    monthly = pnl_data.get("monthly", [])
    annual = pnl_data.get("annual", {})

    row = 2
    for m in monthly:
        idx = m.get("month", 0) - 1
        ws.cell(row=row, column=1, value=months_es[idx] if 0 <= idx < 12 else str(m.get("month", "")))
        ws.cell(row=row, column=2, value=m.get("dividends_gross", 0)).number_format = '"$"#,##0'
        ws.cell(row=row, column=3, value=m.get("dividends_net", 0)).number_format = '"$"#,##0'
        ws.cell(row=row, column=4, value=m.get("wht", 0)).number_format = '"$"#,##0'
        ws.cell(row=row, column=5, value=m.get("options_closed_pnl", 0)).number_format = '"$"#,##0;[Red]"-$"#,##0'
        ws.cell(row=row, column=6, value=m.get("stocks_realized_pnl", 0)).number_format = '"$"#,##0;[Red]"-$"#,##0'
        ws.cell(row=row, column=7, value=m.get("total_income", 0)).number_format = '"$"#,##0;[Red]"-$"#,##0'
        ws.cell(row=row, column=7).font = Font(bold=True, name="SF Mono", size=10, color="30D158" if (m.get("total_income", 0) or 0) >= 0 else "FF453A")
        row += 1

    # Annual total
    ws.cell(row=row, column=1, value="ANUAL").font = Font(bold=True, color="C8A44E")
    ws.cell(row=row, column=2, value=annual.get("dividends_gross", 0)).number_format = '"$"#,##0'
    ws.cell(row=row, column=3, value=annual.get("dividends_net", 0)).number_format = '"$"#,##0'
    ws.cell(row=row, column=4, value=annual.get("wht", 0)).number_format = '"$"#,##0'
    ws.cell(row=row, column=5, value=annual.get("options_closed_pnl", 0)).number_format = '"$"#,##0;[Red]"-$"#,##0'
    ws.cell(row=row, column=6, value=annual.get("stocks_realized_pnl", 0)).number_format = '"$"#,##0;[Red]"-$"#,##0'
    ws.cell(row=row, column=7, value=annual.get("total_income", 0)).number_format = '"$"#,##0;[Red]"-$"#,##0'
    ws.cell(row=row, column=7).font = Font(bold=True, name="SF Mono", size=11, color="C8A44E")

    ws.freeze_panes = "A2"
    return annual.get("total_income", 0)


def dividendos_sheet(wb, divs_data):
    """Dividendos por año (extraído de pnl/monthly byYear si está disponible)."""
    ws = wb.create_sheet("Dividendos")
    headers = ["Año", "Bruto USD", "Neto USD", "WHT USD", "Total Income"]
    widths = [10, 14, 14, 12, 14]
    write_header(ws, 1, headers, widths)

    # divs_data viene de pnl/monthly byYear[] (más confiable que /api/dividendos/resumen)
    by_year = divs_data if isinstance(divs_data, list) else []
    row = 2
    for yr in sorted(by_year, key=lambda x: -(x.get("year", 0) or 0)):
        ws.cell(row=row, column=1, value=yr.get("year", ""))
        ws.cell(row=row, column=2, value=yr.get("dividends_gross", 0) or 0).number_format = '"$"#,##0'
        ws.cell(row=row, column=3, value=yr.get("dividends_net", 0) or 0).number_format = '"$"#,##0'
        ws.cell(row=row, column=4, value=yr.get("wht", 0) or 0).number_format = '"$"#,##0'
        ws.cell(row=row, column=5, value=yr.get("total_income", 0) or 0).number_format = '"$"#,##0;[Red]"-$"#,##0'
        row += 1
    if row == 2:
        ws.cell(row=2, column=1, value="(sin datos byYear disponibles)").font = Font(italic=True, color="8E8E93")

    ws.freeze_panes = "A2"


def earnings_updates_sheet(wb, items):
    ws = wb.create_sheet("Earnings Updates")
    headers = ["Ticker", "Date", "ID", "Size (KB)", "Created"]
    widths = [10, 12, 6, 10, 20]
    write_header(ws, 1, headers, widths)

    row = 2
    for it in items:
        ws.cell(row=row, column=1, value=it.get("ticker", "")).font = Font(bold=True, name="SF Mono", size=10)
        ws.cell(row=row, column=2, value=it.get("date", ""))
        ws.cell(row=row, column=3, value=it.get("id", 0))
        ws.cell(row=row, column=4, value=round((it.get("size_bytes", 0) or 0) / 1024, 1))
        ws.cell(row=row, column=5, value=it.get("created_at", "")[:16].replace("T", " "))
        row += 1

    if row == 2:
        ws.cell(row=2, column=1, value="(sin earnings updates aún)").font = Font(italic=True, color="8E8E93")

    ws.freeze_panes = "A2"


def theses_sheet(wb, theses):
    ws = wb.create_sheet("Tesis")
    headers = ["Ticker", "Tipo", "Conviction", "Target weight", "Updated", "Why owned (resumen)"]
    widths = [10, 12, 12, 14, 12, 80]
    write_header(ws, 1, headers, widths)

    row = 2
    for t in theses:
        ws.cell(row=row, column=1, value=t.get("ticker", "")).font = Font(bold=True, name="SF Mono", size=10)
        ws.cell(row=row, column=2, value=t.get("thesis_type", "—"))
        conv = t.get("conviction", 0)
        ws.cell(row=row, column=3, value=f"{'⭐' * conv if conv else '—'}")
        wmin = t.get("target_weight_min", 0) or 0
        wmax = t.get("target_weight_max", 0) or 0
        ws.cell(row=row, column=4, value=f"{wmin}-{wmax}%" if wmax > 0 else "—")
        ws.cell(row=row, column=5, value=t.get("updated_at", "")[:10] if t.get("updated_at") else "—")
        why = (t.get("why_owned") or "").replace("\n", " ")[:200] + ("..." if len(t.get("why_owned") or "") > 200 else "")
        ws.cell(row=row, column=6, value=why).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 40
        row += 1

    ws.freeze_panes = "A2"


def cover_kpis(ws, total_value, position_count, annual_pnl, snapshot_date, num_updates, num_theses):
    """Llena los KPIs del cover sheet."""
    row = 6
    kpis = [
        ("Cartera (USD)", f"${total_value:,.0f}"),
        ("Posiciones activas", f"{position_count}"),
        ("Income realizado YTD", f"${annual_pnl:,.0f}"),
        ("Earnings Updates totales", f"{num_updates}"),
        ("Tesis escritas", f"{num_theses}"),
    ]
    for label, value in kpis:
        ws.cell(row=row, column=1, value=label).font = Font(size=10, color="8E8E93")
        ws.cell(row=row, column=2, value=value).font = Font(bold=True, size=14, name="SF Mono")
        row += 1

    # Sección sheets disponibles
    ws.cell(row=row + 1, column=1, value="SHEETS").font = Font(bold=True, size=12)
    sheets_info = [
        ("Cartera", "Posiciones con P&L + yield + weight"),
        ("PnL Mensual", "Dividendos + opciones + stocks por mes"),
        ("Dividendos", "Histórico anual"),
        ("Earnings Updates", "Reports generados Claude Code"),
        ("Tesis", "Investment thesis por position"),
    ]
    for s_name, s_desc in sheets_info:
        ws.cell(row=row + 2, column=1, value=s_name).font = Font(bold=True, color="C8A44E")
        ws.cell(row=row + 2, column=2, value=s_desc).font = Font(size=10, color="8E8E93")
        row += 1

    # Footer
    ws.cell(row=row + 4, column=1, value="🤖 Generado coste $0 vía Claude Code subscription.").font = Font(italic=True, color="8E8E93", size=9)
    ws.cell(row=row + 5, column=1, value="📁 Fuente datos: A&R API (positions, pnl/monthly, dividendos, earnings_updates, theses)").font = Font(italic=True, color="8E8E93", size=9)


def main():
    parser = argparse.ArgumentParser(description="A&R Monthly Snapshot to xlsx")
    parser.add_argument("--month", help="YYYY-MM format (default: actual)")
    parser.add_argument("--output", help="Override output path")
    args = parser.parse_args()

    if args.month:
        try:
            snapshot_date = datetime.strptime(args.month + "-01", "%Y-%m-%d")
        except ValueError:
            print(f"ERROR: --month must be YYYY-MM format")
            sys.exit(1)
    else:
        snapshot_date = datetime.now()

    year = snapshot_date.year
    month_str = snapshot_date.strftime("%Y-%m")

    # Output path
    output = Path(args.output) if args.output else ICLOUD_DIR / f"{month_str}.xlsx"
    output.parent.mkdir(parents=True, exist_ok=True)

    print(f"📊 A&R Monthly Snapshot — {month_str}")
    print(f"🔌 Fetching data from {API_BASE}...")

    # Fetch all data
    print("  · positions")
    positions = fetch("/api/positions").get("positions", [])
    print("  · pnl/monthly")
    pnl = fetch(f"/api/pnl/monthly?year={year}&broker=ALL")
    # Use pnl/monthly byYear[] como fuente confiable (auth-friendly)
    divs = pnl.get("byYear", []) if isinstance(pnl, dict) else []
    print("  · earnings/auto-update/list")
    eu = fetch("/api/earnings/auto-update/list").get("items", [])
    print("  · theses")
    theses = fetch("/api/theses").get("theses", [])

    # Build workbook
    wb = Workbook()
    wb.remove(wb.active)  # remove default

    cover = cover_sheet(wb, snapshot_date)
    total_value, position_count = positions_sheet(wb, positions)
    annual_pnl = pnl_sheet(wb, pnl)
    dividendos_sheet(wb, divs)
    earnings_updates_sheet(wb, eu)
    theses_sheet(wb, theses)
    cover_kpis(cover, total_value, position_count, annual_pnl, snapshot_date, len(eu), len(theses))

    wb.save(output)
    size_kb = output.stat().st_size / 1024
    print(f"\n✓ Saved: {output}")
    print(f"  Size: {size_kb:.1f} KB · Sheets: {len(wb.sheetnames)}")
    print(f"  Cartera ${total_value:,.0f} ({position_count} pos)")
    print(f"  YTD P&L: ${annual_pnl:,.0f}")


if __name__ == "__main__":
    main()
