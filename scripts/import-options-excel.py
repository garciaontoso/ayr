#!/usr/bin/env python3
"""
import-options-excel.py — One-shot importer for the A&R master Excel.

Reads the Credit Spreads / ROC / ROP sheets for years 2023-2026 and uploads
every trade to the worker via POST /api/options/trades/bulk-import. Handles:

  - Sheet layouts: C.S (with long strike + Kelly) / ROC (covered calls) /
    ROP (cash-secured puts). Row positions are identical across 23-26.
  - Status normalisation: "CLOSED " → "CLOSED", etc.
  - Account detection: cells below the trade area ("IB" or "TASTY")
  - IDEA rows are skipped (not executed trades)
  - Formula-derived values are read via data_only=True (LibreOffice needed
    if values are not cached; otherwise openpyxl returns None).
  - Issues are logged to options_import_issues via the bulk endpoint.

Usage:
  source ~/.ayr-env && scripts/import-options-excel.py [--dry-run] [--sheet 'C.S 26'] [--file PATH]
"""
import argparse
import json
import os
import sys
import time
import urllib.error
import urllib.request
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip3 install openpyxl")
    sys.exit(1)

WORKER_URL = os.environ.get("AYR_WORKER_URL", "https://aar-api.garciaontoso.workers.dev")
WORKER_TOKEN = os.environ.get("AYR_WORKER_TOKEN", "")
DEFAULT_FILE = "/Users/ricardogarciaontoso/Downloads/1- A $ R (3).xlsx"


# ─── Layout definitions ───────────────────────────────────────────────
# Map of row-number → field name. Rows are IDENTICAL across years 23-26.
CS_ROWS = {
    3: 'trade_date', 4: 'underlying', 5: 'price', 6: 'on_sale_price',
    7: 'dte', 8: 'expiration_date', 9: 'floor_ceiling', 10: 'buffer_pct',
    11: 'floor_buffer_strike', 12: 'actual_pct_from_floor', 13: 'prob_otm',
    14: 'delta', 15: 'short_strike', 16: 'actual_pct_from_price',
    17: 'adj_pct_from_price', 18: 'long_strike', 19: 'spread',
    20: 'target_credit', 21: 'credit', 22: 'commission', 23: 'net_credit',
    24: 'risk_capital', 25: 'margin_pct', 26: 'margin_capital', 27: 'rorc',
    28: 'multiplier', 29: 'arorc', 30: 'qtr_report_flag',
    32: 'kelly_w', 33: 'rc_at_risk_pct', 34: 'avg_loss', 35: 'kelly_r',
    36: 'kelly_pct', 37: 'bankroll', 38: 'kelly_max_bet',
    39: 'rule1_max_margin', 40: 'max_contracts', 41: 'actual_contracts',
    42: 'shares', 43: 'net_credit_total', 44: 'risk_capital_total',
    57: 'status', 58: 'result_date', 59: 'closing_debit', 60: 'total_debit',
    61: 'final_net_credit', 62: 'final_rorc', 63: 'final_arorc', 64: 'notes',
}

ROC_ROWS = {
    3: 'trade_date', 4: 'underlying', 5: 'price', 6: 'on_sale_price',
    7: 'dte', 8: 'expiration_date', 9: 'floor_ceiling', 10: 'buffer_pct',
    11: 'floor_buffer_strike', 12: 'actual_pct_from_floor', 13: 'prob_otm',
    14: 'delta', 15: 'short_strike', 16: 'actual_pct_from_price',
    17: 'target_credit', 18: 'credit', 19: 'commission', 20: 'net_credit',
    21: 'risk_capital', 22: 'margin_capital', 23: 'rorc', 24: 'multiplier',
    25: 'arorc', 26: 'qtr_report_flag', 27: 'bankroll', 28: 'actual_contracts',
    29: 'shares', 30: 'net_credit_total', 31: 'risk_capital_total',
    44: 'status', 45: 'result_date', 46: 'closing_debit', 47: 'total_debit',
    48: 'final_net_credit', 49: 'final_rorc', 50: 'final_arorc', 51: 'notes',
}

ROP_ROWS = {
    3: 'trade_date', 4: 'underlying', 5: 'price', 6: 'on_sale_price',
    7: 'dte', 8: 'expiration_date', 9: 'floor_ceiling', 10: 'buffer_pct',
    11: 'floor_buffer_strike', 12: 'actual_pct_from_floor', 13: 'prob_otm',
    14: 'delta', 15: 'short_strike', 16: 'actual_pct_from_price',
    17: 'adj_pct_from_price', 18: 'target_credit', 19: 'credit',
    20: 'commission', 21: 'net_credit', 22: 'risk_capital', 23: 'margin_pct',
    24: 'margin_capital', 25: 'rorc', 26: 'multiplier', 27: 'arorc',
    28: 'qtr_report_flag', 29: 'kelly_w', 30: 'bankroll',
    31: 'actual_contracts', 32: 'shares', 33: 'net_credit_total',
    34: 'risk_capital_total',
    45: 'status', 46: 'result_date', 47: 'closing_debit', 48: 'total_debit',
    49: 'final_net_credit', 50: 'final_rorc', 51: 'final_arorc', 52: 'notes',
}

# 2024 ROP has RESULT at row 44 (not 45 like 25/26 — extra row shift).
# We'll probe both rows and pick the one that has a plausible status.
ROP_24_ALT_ROWS = {**ROP_ROWS}
ROP_24_ALT_ROWS.pop(45); ROP_24_ALT_ROWS[44] = 'status'
ROP_24_ALT_ROWS.pop(46); ROP_24_ALT_ROWS[45] = 'result_date'
ROP_24_ALT_ROWS.pop(47); ROP_24_ALT_ROWS[46] = 'closing_debit'
ROP_24_ALT_ROWS.pop(48); ROP_24_ALT_ROWS[47] = 'total_debit'
ROP_24_ALT_ROWS.pop(49); ROP_24_ALT_ROWS[48] = 'final_net_credit'
ROP_24_ALT_ROWS.pop(50); ROP_24_ALT_ROWS[49] = 'final_rorc'
ROP_24_ALT_ROWS.pop(51); ROP_24_ALT_ROWS[50] = 'final_arorc'
ROP_24_ALT_ROWS.pop(52); ROP_24_ALT_ROWS[51] = 'notes'

SHEETS = [
    ('C.S 26', 'CS', 2026, CS_ROWS),
    ('C.S 25', 'CS', 2025, CS_ROWS),
    ('C.S 24', 'CS', 2024, CS_ROWS),
    ('C.S 23', 'CS', 2023, CS_ROWS),
    ('ROC 26', 'ROC', 2026, ROC_ROWS),
    ('ROC 25', 'ROC', 2025, ROC_ROWS),
    ('ROC 24', 'ROC', 2024, ROC_ROWS),
    ('ROC 23', 'ROC', 2023, ROC_ROWS),
    ('ROP 26', 'ROP', 2026, ROP_ROWS),
    ('ROP 25', 'ROP', 2025, ROP_ROWS),
    ('ROP 24', 'ROP', 2024, ROP_24_ALT_ROWS),
    ('ROP 23', 'ROP', 2023, ROP_24_ALT_ROWS),
]

VALID_STATUSES = {'OPEN', 'EXPIRED', 'CLOSED', 'ROLLED', 'ASSIGNED', 'IDEA'}


# ─── Value coercion ────────────────────────────────────────────────────
def coerce(field, v):
    if v is None:
        return None
    if isinstance(v, datetime):
        if field in ('trade_date', 'expiration_date', 'result_date'):
            return v.date().isoformat()
        return v.isoformat()
    if isinstance(v, str):
        s = v.strip()
        if not s or s.lower() in ('n/a', 'na', '-', '—', '#n/a'):
            return None
        if field == 'status':
            u = s.upper().replace(' ', '')
            if u in VALID_STATUSES:
                return u
            return s.upper()
        if field in ('qtr_report_flag', 'underlying', 'notes', 'account'):
            return s
        # Try numeric
        try:
            return float(s.replace(',', '').replace('$', '').replace('%', ''))
        except ValueError:
            return s
    if isinstance(v, (int, float)):
        if field == 'status':
            return None  # numbers in status = noise
        return float(v)
    return v


def extract_trade(ws, col, row_map, strategy, year, sheet_name, issues):
    """Extract one trade from a column. Returns dict or None to skip."""
    t = {
        'strategy': strategy,
        'year': year,
        'source_sheet': sheet_name,
        'source_col': col,
    }
    for row, field in row_map.items():
        raw = ws.cell(row, col).value
        val = coerce(field, raw)
        if val is not None:
            t[field] = val

    # Must have underlying and either trade_date or expiration_date
    if not t.get('underlying') or (not t.get('trade_date') and not t.get('expiration_date')):
        return None

    # Skip IDEA trades
    if t.get('status') == 'IDEA':
        return None

    # Default status to OPEN if missing
    if not t.get('status'):
        t['status'] = 'OPEN'

    # Auto-detect account: look at rows 65-95 in same column for IB/TASTY markers
    for probe_row in range(65, 100):
        pv = ws.cell(probe_row, col).value
        if isinstance(pv, str):
            pu = pv.strip().upper()
            if pu == 'IB':
                t['account'] = 'IB'
                break
            if pu.startswith('TASTY'):
                t['account'] = 'TASTY'
                break
    if not t.get('account'):
        t['account'] = 'IB'  # default

    # Sanity checks → log as issues, don't block
    if t.get('status') in ('CLOSED', 'EXPIRED', 'ASSIGNED', 'ROLLED') and not t.get('result_date'):
        issues.append({
            'source_sheet': sheet_name, 'source_col': col,
            'severity': 'warning', 'category': 'missing_result_date',
            'message': f"{strategy} trade for {t.get('underlying')} is {t['status']} but has no result_date",
        })

    if strategy == 'CS' and not t.get('long_strike'):
        issues.append({
            'source_sheet': sheet_name, 'source_col': col,
            'severity': 'warning', 'category': 'missing_long_strike',
            'message': f"CS trade for {t.get('underlying')} at col {col} has no long_strike",
        })

    return t


def post_batch(trades, issues, dry_run=False):
    if dry_run:
        print(f"  [dry-run] Would POST {len(trades)} trades + {len(issues)} issues")
        return {'inserted': 0, 'updated': 0, 'skipped': 0}
    if not WORKER_TOKEN:
        print("ERROR: AYR_WORKER_TOKEN not set. Source ~/.ayr-env first.")
        sys.exit(1)
    payload = json.dumps({'trades': trades, 'issues': issues}).encode('utf-8')
    req = urllib.request.Request(
        f"{WORKER_URL}/api/options/trades/bulk-import",
        data=payload, method='POST'
    )
    req.add_header('Content-Type', 'application/json')
    req.add_header('Authorization', f'Bearer {WORKER_TOKEN}')
    req.add_header('User-Agent', 'Mozilla/5.0 (Macintosh; AyR Options Importer/1.0)')
    try:
        with urllib.request.urlopen(req, timeout=120) as resp:
            return json.loads(resp.read().decode('utf-8'))
    except urllib.error.HTTPError as e:
        body = e.read().decode('utf-8', errors='replace')[:500]
        raise Exception(f"HTTP {e.code}: {body}")


def process_sheet(wb, sheet_name, strategy, year, row_map, dry_run=False):
    if sheet_name not in wb.sheetnames:
        print(f"[{sheet_name}] NOT FOUND, skipping")
        return 0, 0, 0
    ws = wb[sheet_name]

    trade_cols = []
    for c in range(3, ws.max_column + 1):
        v = ws.cell(4, c).value  # UNDERLYING row
        if isinstance(v, str) and v.strip() and v.upper() not in (
            'UNDERLYING', 'COMPRA ACCIONES', 'SEMANAL', 'MENSUAL',
            'TOTAL IB', 'TOTAL TASTY', 'TOTAL'
        ):
            trade_cols.append(c)

    trades = []
    issues = []
    for c in trade_cols:
        t = extract_trade(ws, c, row_map, strategy, year, sheet_name, issues)
        if t:
            trades.append(t)

    print(f"[{sheet_name}] {len(trade_cols)} cols scanned, {len(trades)} trades extracted, {len(issues)} issues")

    # Upload in batches of 50
    inserted = updated = skipped = 0
    batch_size = 50
    for i in range(0, len(trades), batch_size):
        batch = trades[i:i+batch_size]
        batch_issues = issues if i == 0 else []  # send issues only on first batch
        try:
            r = post_batch(batch, batch_issues, dry_run=dry_run)
            inserted += r.get('inserted', 0)
            updated += r.get('updated', 0)
            skipped += r.get('skipped', 0)
            print(f"  batch {i//batch_size + 1}: +{r.get('inserted',0)} ~{r.get('updated',0)} !{r.get('skipped',0)}")
        except Exception as e:
            print(f"  batch {i//batch_size + 1} FAILED: {e}")
            skipped += len(batch)
        time.sleep(0.3)  # gentle pace
    return inserted, updated, skipped


def main():
    p = argparse.ArgumentParser()
    p.add_argument('--file', default=DEFAULT_FILE, help='Path to the Excel file')
    p.add_argument('--sheet', help='Import only one sheet (e.g. "C.S 26")')
    p.add_argument('--dry-run', action='store_true')
    args = p.parse_args()

    if not Path(args.file).exists():
        print(f"ERROR: file not found: {args.file}")
        sys.exit(1)

    print(f"Loading {args.file} (data_only=True)...")
    wb = load_workbook(args.file, data_only=True)
    print(f"  {len(wb.sheetnames)} sheets")

    total_ins = total_upd = total_skp = 0
    sheets_to_process = [s for s in SHEETS if not args.sheet or s[0] == args.sheet]
    for sheet_name, strategy, year, row_map in sheets_to_process:
        i, u, s = process_sheet(wb, sheet_name, strategy, year, row_map, dry_run=args.dry_run)
        total_ins += i; total_upd += u; total_skp += s

    print(f"\n=== DONE ===")
    print(f"Inserted: {total_ins}")
    print(f"Updated:  {total_upd}")
    print(f"Skipped:  {total_skp}")
    print(f"Total:    {total_ins + total_upd + total_skp}")


if __name__ == '__main__':
    main()
